
from __future__ import annotations

from pathlib import Path
from hashlib import sha1
from typing import Dict, List
from datetime import datetime, date
import math

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =========================================================
# CONFIGURAÇÕES
# =========================================================
ARQUIVO_CARTEIRA = Path(r"Z:\Asset Management\Equipe\Lívia\Rascunhos\Comparação Fundo de Terceiros\Fundo-de-Terceiros\carteira_fundos_consolidada.xlsx")
ARQUIVO_SAIDA = ARQUIVO_CARTEIRA
ARQUIVO_BASE_MANUAL = ARQUIVO_CARTEIRA.with_name("cadastro_manual_ativos.xlsx")

ABA_BASE_MANUAL = "base_ativos"
ABA_CONTROLE_CHAVES = "_controle_chaves"

COLUNAS_CONTROLE_CHAVES = [
    "Id_Ativo",
    "Chave_Original",
    "Chave_Final",
]

COLUNAS_FIXAS = [
    "Codigo",
    "ISIN",
    "Descrição",
    "Emissor",
    "Classe",
    "Tipo de Investimento",
    "Vencimento",
]

COLUNAS_MANUAIS = [f"{col} - Preencher" for col in COLUNAS_FIXAS]
COLUNAS_BASE_MANUAL = [
    "Id_Ativo",
    "Campos_Faltantes",
] + COLUNAS_FIXAS + COLUNAS_MANUAIS

MARCADORES_FIM = {"SOMA_COLUNA", "PL_CDA_FI_PL", "DIFERENCA", "STATUS"}

# False = o manual pode sobrescrever o valor da carteira
# True = só preenche campo vazio
PREENCHER_APENAS_VAZIOS = False


# =========================================================
# HELPERS
# =========================================================
def normalizar_texto(valor) -> str:
    if valor is None:
        return "-"

    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")

    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")

    if isinstance(valor, float):
        if math.isnan(valor):
            return "-"
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()

    texto = str(valor).strip()
    return texto if texto else "-"


def esta_vazio(valor) -> bool:
    return normalizar_texto(valor) in {"-", "", "None", "nan", "NaN"}


def formatar_valor_para_excel(coluna: str, valor):
    if esta_vazio(valor):
        return "-"

    if coluna == "Vencimento":
        if isinstance(valor, datetime):
            return valor.strftime("%Y-%m-%d")
        if isinstance(valor, date):
            return valor.strftime("%Y-%m-%d")

        texto = str(valor).strip()
        try:
            return datetime.fromisoformat(texto).strftime("%Y-%m-%d")
        except Exception:
            pass

        for fmt in (
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y/%m/%d",
            "%Y-%m-%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%Y-%m-%d",
        ):
            try:
                return datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
            except Exception:
                continue

        return texto

    return str(valor).strip()


def eh_linha_totalmente_vazia(ws, row_idx: int) -> bool:
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=row_idx, column=col_idx).value not in (None, ""):
            return False
    return True


def localizar_colunas(ws) -> Dict[str, int]:
    mapa: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        cabecalho = ws.cell(row=1, column=col_idx).value
        if cabecalho is not None:
            mapa[str(cabecalho).strip()] = col_idx
    return mapa


def iterar_linhas_dados(ws):
    for row_idx in range(2, ws.max_row + 1):
        valor_a = normalizar_texto(ws.cell(row=row_idx, column=1).value)

        if valor_a in MARCADORES_FIM:
            break

        if eh_linha_totalmente_vazia(ws, row_idx):
            continue

        yield row_idx


def gerar_chave_exata_ativo(valores_fixos: Dict[str, object]) -> str:
    return "||".join(normalizar_texto(valores_fixos.get(col)) for col in COLUNAS_FIXAS)


def gerar_id_ativo(valores_fixos: Dict[str, object]) -> str:
    texto = gerar_chave_exata_ativo(valores_fixos)
    return sha1(texto.encode("utf-8")).hexdigest()[:20]


def gerar_id_por_chave(chave: str) -> str:
    return sha1(chave.encode("utf-8")).hexdigest()[:20]


def gerar_chave_final_manual(registro: Dict[str, object], manual: Dict[str, object]) -> str:
    valores_finais = {}
    for col in COLUNAS_FIXAS:
        valor_manual = manual.get(col)
        if not esta_vazio(valor_manual):
            valores_finais[col] = valor_manual
        else:
            valores_finais[col] = registro.get(col)
    return gerar_chave_exata_ativo(valores_finais)


def normalizar_registro_fixos(valores_fixos: Dict[str, object]) -> Dict[str, str]:
    return {col: formatar_valor_para_excel(col, valores_fixos.get(col)) for col in COLUNAS_FIXAS}


def contar_faltantes(valores_fixos: Dict[str, object]) -> List[str]:
    return [col for col in COLUNAS_FIXAS if esta_vazio(valores_fixos.get(col))]

def contar_preenchidos(registro: Dict[str, object]) -> int:
    return sum(0 if esta_vazio(registro.get(col)) else 1 for col in COLUNAS_FIXAS)

def deve_aplicar_valor(valor_atual, novo_valor) -> bool:
    if esta_vazio(novo_valor):
        return False
    if PREENCHER_APENAS_VAZIOS:
        return esta_vazio(valor_atual)
    return normalizar_texto(valor_atual) != normalizar_texto(novo_valor)


def to_float_seguro(valor) -> float:
    if valor is None:
        return 0.0

    if isinstance(valor, (int, float)):
        try:
            if isinstance(valor, float) and math.isnan(valor):
                return 0.0
        except Exception:
            pass
        return float(valor)

    texto = str(valor).strip()
    if not texto or texto == "-":
        return 0.0

    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except Exception:
        return 0.0


def eh_coluna_data(nome_coluna: str) -> bool:
    texto = str(nome_coluna).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            datetime.strptime(texto, fmt)
            return True
        except Exception:
            continue
    return False

PADRONIZACAO_EMISSORES_LF = {
    "AGROLEND SOCIEDADE DE CREDITO FINAN E IN": "Agrolend Sociedade de Crédito",
    "BANCO ABC BRASIL S.A.": "Banco ABC Brasil",
    "BANCO AGIBANK S": "Banco Agibank",
    "BANCO AGIBANK S.A": "Banco Agibank",
    "BANCO AGIBANK S.A.": "Banco Agibank",
    "BANCO AGIPLAN S.A.": "Banco Agiplan",
    "BANCO ALFA DE INVESTIMENTO S.A.": "Banco Alfa",
    "BANCO ALFA DE INVESTIMENTOS S/A": "Banco Alfa",
    "BANCO BMG S.A": "Banco BMG",
    "BANCO BMG SA": "Banco BMG",
    "BANCO BOCOM BBM S.A.": "Banco BOCOM BBM",
    "BANCO BRADESCO": "Banco Bradesco",
    "BANCO BRADESCO S.A": "Banco Bradesco",
    "BANCO BRADESCO S.A.": "Banco Bradesco",
    "BANCO BTG PACTUAL S A": "Banco BTG Pactual",
    "BANCO BTG PACTUAL S.A.": "Banco BTG Pactual",
    "BANCO BTG PACTUAL S/A": "Banco BTG Pactual",
    "BANCO C6 S.A.": "Banco C6",
    "BANCO CAIXA GERAL - BRASIL S.A.": "Banco Caixa",
    "BANCO CITIBANK S A": "Banco Citibank",
    "BANCO CITIBANK S.A.": "Banco Citibank",
    "BANCO CNH CAPITAL S.A.": "Banco CNH Capital",
    "BANCO CNH CAPIT": "Banco CNH Capital",
    "BANCO CNH INDUSTRIAL CAPITAL S.A.": "Banco CNH Industrial",
    "BANCO COOPERATI": "Banco Sicred",
    "BANCO COOPERATIVO SICREDI S.A.": "Banco Sicred",
    "BANCO C6 CONSIGNADO S.A.": "Banco C6 Consignado",
    "BANCO CSF S/A": "Banco CSF",
    "BANCO DAYCOVAL": "Banco Daycoval",
    "BANCO DAYCOVAL S.A.": "Banco Daycoval",
    "BANCO DE LAGE LANDEN BRASIL S.A.": "Banco de Lage Landen",
    "BANCO DO BRASIL": "Banco do Brasil",
    "BANCO DO BRASIL SA": "Banco do Brasil",
    "BANCO DO ESTADO DO RIO GRANDE DO SUL SA": "Banco do Estado do Rio Grande do Sul",
    "BANCO DO NORDESTE DO BRASIL SA": "Banco do Nordeste",
    "BANCO FIDIS DE INVESTIMENTO S.A.": "Banco FIDIS de Investimento",
    "BANCO FIDIS S.A": "Banco FIDIS de Investimento",
    "BANCO GM S.A": "Banco GM",
    "BANCO GMAC S.A": "Banco GMAC",
    "BANCO GMAC SA": "Banco GMAC",
    "BANCO HONDA S/A": "Banco Honda",
    "BANCO HSBC SA": "Banco HSBC",
    "BANCO IBM S/A": "Banco IBM",
    "BANCO INDUSTRIAL DO BRASIL S/A": "Banco Industrial do Brasil",
    "BANCO  INDUSTRIAL BRASIL SA": "Banco Industrial do Brasil",
    "BANCO INTER SA": "Banco Inter",
    "BANCO INTER S.A": "Banco Inter",
    "BANCO INTERMEDI": "Banco Intermedium",
    "BANCO INTERMEDIUM S.A.": "Banco Intermedium",
    "BANCO ITAU BBA S.A.": "Banco Itaú",
    "BANCO JOHN DEERE S.A.": "Banco John Deere",
    "BANCO MERCANTIL DO BRASIL S.A.": "Banco Mercantil",
    "BANCO MERCANTIL DO BRASIL S/A": "Banco Mercantil",
    "BANCO MERCANTIL DO BRASIL SA": "Banco Mercantil",
    "BANCO MERCEDES": "Banco Mercedes",
    "BANCO MERCEDES-BENZ DO BRASIL S.A.": "Banco Mercedes",
    "BANCO MERCEDES-BENZ DO BRASIL S/A": "Banco Mercedes",
    "BANCO ORIGINAL S.A": "Banco Original",
    "BANCO PACCAR SA": "Banco Paccar",
    "BANCO PAN S.A": "Banco Pan",
    "BANCO PAN S.A.": "Banco Pan",
    "BANCO PSA FINAN": "Banco PSA Finance",
    "BANCO PSA FINANCE BRASIL S.A.": "Banco PSA Finance",
    "BANCO RANDON S/A": "Banco Randon",
    "BANCO RANDON SA": "Banco Randon",
    "BANCO RCI BRASIL S.A": "Banco RCI Brasil",
    "BANCO RCI BRASIL S.A.": "Banco RCI Brasil",
    "BANCO RODOBENS": "Banco Rodobens",
    "BANCO RODOBENS S.A.": "Banco Rodobens",
    "BANCO SAFRA S A": "Banco Safra",
    "BANCO SAFRA S.A": "Banco Safra",
    "BANCO SAFRA S.A.": "Banco Safra",
    "BANCO SAFRA S/A": "Banco Safra",
    "BANCO SANTANDER": "Banco Santander",
    "BANCO SANTANDER (BRASIL) S.A.": "Banco Santander",
    "BANCO SANTANDER (BRASIL) S/A.": "Banco Santander",
    "BANCO SEGURO S.A.": "Banco Seguro",
    "BANCOSEGURO S.A": "Banco Seguro",  
    "BANCO SOFISA": "Banco Sofisa",
    "BANCO SOFISA S.A.": "Banco Sofisa",
    "BANCO SOFISA S/A": "Banco Sofisa",
    "BANCO SOFISA SA": "Banco Sofisa",
    "BANCO STELLANTIS S.A.": "Banco Stellantis",
    "BANCO SUMITOMO MITSUI BRASILEIRO SA": "Banco Sumitomo Mitsui",
    "BANCO TOYOTA DO BRASIL S.A.": "Banco Toyota",
    "BANCO VOLKSWAGE": "Banco Volkswagen",
    "BANCO VOLKSWAGEN S.A.": "Banco Volkswagen",
    "BANCO VOLKSWAGEN S/A": "Banco Volkswagen",
    "BANCO VOLKSWAGEN SA": "Banco Volkswagen",
    "BANCO VOTORANTIM S.A.": "Banco Votorantim",
    "BANCO VOTORANTIM S/A": "Banco Votorantim",
    "BANCO XP S.A": "Banco XP",
    "BANCO XP S/A": "Banco XP",
    "BANCO YAMAHA MOTOR BRASIL S.A.": "Banco Yamaha",
    "BANCO YAMAHA MOTOR DO BRASIL SA": "Banco Yamaha",
    "BANCOSEGURO SA": "Banco Seguro",
    "BANRISUL": "Banco Banrisul",
    "BCO ABC BRASIL": "Banco ABC Brasil",
    "BCO ABC BRASIL SA": "Banco ABC Brasil",
    "BCO ABC BRASIL SA (EX BCO ABC ROMA SA)": "Banco ABC Brasil",
    "BCO BMG SA": "Banco BMG",
    "BCO BRADESCO SA": "Banco Bradesco",
    "BCO BRASIL SA": "Banco do Brasil",
    "BCO BTG PACTUAL": "Banco BTG Pactual",
    "BCO BTG PACTUAL SA": "Banco BTG Pactual",
    "BCO CAIXA GERAL BRASIL SA": "Banco Caixa",
    "BCO CITIBANK SA": "Banco Citibank",
    "BCO CNH CAPITAL SA": "Banco CNH Capital",
    "BCO COOPERATIVO SICREDI SA  BANSICREDI": "Banco Sicred",
    "BCO DAYCOVAL SA": "Banco Daycoval",
    "BCO ESTADO DO RIO GRANDE DO SUL S.A": "Banco do Estado do Rio Grande do Sul",
    "BCO ESTADO PARA": "Banco do Estado Para",
    "BCO ESTADO PARA SA": "Banco do Estado Para",
    "BCO ESTADO RIO GRANDE SUL SA": "Banco do Estado do Rio Grande do Sul",
    "BCO GMAC SA": "Banco GMAC",
    "BCO INDUSTRIAL": "Banco Industrial do Brasil",
    "BCO JOHN DEERE SA": "Banco John Deere",
    "BCO MERCANTIL B": "Banco Mercantil",
    "BCO MERCANTIL BRASIL SA": "Banco Mercantil",
    "BCO MERCANTIL DO BRASIL S/A": "Banco Mercantil",
    "BCO MERCEDES BENZ BRASIL SA": "Banco Mercedes",
    "BCO NORDESTE BRASIL SA": "Banco do Nordeste",
    "BCO ORIGINAL AGRONEG - Ex JBS BCO S": "Banco Original",
    "BCO RABOBANK INTERNATIONAL BRASIL": "Banco Rabobank International Brasil",
    "BCO REGIONAL DESENVOLVIMENTO EXTREMO SUL BRDE": "Banco Regional Desenvolvimento Extremo Sul",
    "BCO RODOBENS S.A": "Banco Rodobens",
    "BCO RODOBENS SA": "Banco Rodobens",
    "BCO SAFRA SA": "Banco Safra",
    "BCO SANTANDER (BRASIL) SA": "Banco Santander",
    "BCO SOFISA SA": "Banco Sofisa",
    "BCO VOLKSWAGEN SA": "Banco Volkswagen",
    "BCO VOTORANTIM": "Banco Votorantim",
    "BR PARTNERS BANCO DE INVESTIMENTO S.A.": "Banco BR Partners",
    "BR PARTNERS BANCO INVESTIMENTO SA": "Banco BR Partners",
    "BR PARTNERS BCO": "Banco BR Partners",
    "Banco  Volkswagen S/A": "Banco Volkswagen",
    "CAIXA ECONOMICA FEDERAL": "Banco Caixa",
    "CAIXA ECONOMICA FEDERAL - CEF": "Banco Caixa",
    "CAIXA ECONOMICA FEDERAL CEF": "Banco Caixa",
    "CEF": "Banco Caixa",
    "CIA DE CR FINAN E INVEST RCI BRASIL": "Banco RCI",
    "CITIBANK": "Banco Citibank",
    "CITIBANK NA FIL": "Banco Citibank",
    "CONCORDIA BANCO S.A.": "Banco Concórdia",
    "COOPERATIVA DE CRÉDITO SICREDI RECIFE - SICREDI RECIFE": "Banco Sicred Recife",
    "DEUTSCHE": "Banco Deutsche",
    "DEUTSCHE BANK SA - BANCO ALEMAO": "Banco Deutsche",
    "FICSA CCTVM LTDA": "Bando Ficsa",
    "FUNDO INVEST QUOTAS FITVM BNP PARIBAS ACTIVE": "BNP Paribas Active FICFITVM", 
    "GOLDMAN SACHS BRASIL BCO MULTIPLO": "Goldman Sachs Brasil",
    "HAITONG BANCO DE INVESTIMENTO DO BRASIL SA": "Banco Haitong",
    "HAITONG BANCO INVESTIMENTO BRASIL SA": "Banco Haitong",
    "HSBC BANK BRASIL SA - BANCO MULTIPL": "Banco HSBC",
    "ITAU UNIBANCO HLDG SA/KY": "Banco Itaú",
    "ITAU UNIBANCO HOLDING S.A.": "Banco Itaú",
    "ITAU UNIBANCO HOLDING SA": "Banco Itaú",
    "ITAU UNIBANCO S": "Banco Itaú",
    "ITAU UNIBANCO S.A.": "Banco Itaú",
    "ITAUBANCO HOLDI": "Banco Itaú",
    "ITAÚ UNIBANCO HOLDING S.A.": "Banco Itaú",
    "MERCADO CREDITO SO. CR. FINAN. E IN": "Mercado Crédito Sociedade de Crédito, Financimento e Investimento",
    "MERCADO CREDITO SOCIEDADE DE CREDITO FINANCIAMENTO": "Mercado Crédito Sociedade de Crédito, Financimento e Investimento",
    "MERCADO CREDITO SOCIEDADE DE CREDITO FINANCIAMENT": "Mercado Crédito Sociedade de Crédito, Financimento e Investimento",
    "MERCADO CREDITO SOCIEDADE DE CREDITO FINANCIAMENTO E INVESTIMENTO SA": "Mercado Crédito Sociedade de Crédito, Financimento e Investimento",
    "Mercado Credito": "Mercado Crédito Sociedade de Crédito, Financimento e Investimento",
    "MIDWAY S.A. CRÉ": "Midway",
    "MIDWAY S.A.- CREDITO, FINANCIAMENTO": "Midway",
    "NBC BANK BRASIL SA BCO MULTIPLO": "Banco NBC Brasil",
    "NU FINANCEIRA S": "Nubank", 
    "NU FINANCEIRA S.A.": "Nubank",
    "NU FINANCEIRA S.A. - SOCIEDADE DE CREDITO FINANCI": "Nubank",
    "NU FINANCEIRA S.A. SOCIEDADE DE CREDITO FINANCIAMENTO E INVESTIMENTO": "Nubank",
    "OMNI BANCO SA": "Banco OMNI",
    "OMNI SA CRED FINANC INVEST": "Banco OMNI",
    "OMNI SA CREDITO": "Banco OMNI",
    "OMNI S/A CREDITO FINANCIAMENTO E INVESTIMENTO": "Banco OMNI",
    "PARANA BANCO": "Banco Paraná",
    "PARANA BANCO S.A.": "Banco Paraná",
    "PARANA BANCO S/A": "Banco Paraná",
    "PARANA BCO SA": "Banco Paraná",
    "PICPAY BANK - B": "Banco Picpay",
    "PICPAY BANK - BANCO MULTIPLO S.A": "Banco Picpay",
    "PICPAY BANK - BANCO MULTIPLO S.A.": "Banco Picpay",
    "PORTOSEG S/A - CREDITO FINANCIAMENTO E INVESTIMEN": "Porto Seguro",
    "PORTOSEG S/A - CREDITO, FINANCIAMEN": "Porto Seguro",
    "PORTOSEG SA CRE": "Porto Seguro",
    "PORTOSEG SA CRED FINANC INVEST": "Porto Seguro",
    "SCANIA BANCO SA": "Banco Scania",
    "STELLANTIS FINANCIAMENTOS SOCIEDADE DE CREDITO FIN": "Banco Stellantis",
    "STONE SOCIEDADE": "Stone Sociedade de Crédito Direto",
    "STONE SOCIEDADE DE CRÉDITO, FINANCIAMENTO E INVESTIMENTO S.A.": "Stone Sociedade de Crédito Direto",
    "STONE SOCIEDADE DE CREDITO FINANCIAMENTO": "Stone Sociedade de Crédito Direto",
    "STONE SOCIEDADE DE CREDITO FINANCIAMENTO E INVEST": "Stone Sociedade de Crédito Direto",
    "Stone Sociedade de Credito, Financi": "Stone Sociedade de Crédito Direto",
    "VIRGO COMPANHIA DE SECURITIZACAO": "Virgo Companhia de Securitização",
    "VOLKSWAGEN DO BRASIL INDÚSTRIA DE VEÍCULOS AUTOMOTORES LTDA.": "Banco Volkswagen",
    "XP INVESTIMENTOS S/A": "Banco XP",

    # adicione outros aqui
}


def normalizar_emissor_para_regra(emissor) -> str:
    if emissor is None:
        return ""
    texto = str(emissor).strip().upper()
    return " ".join(texto.split())


def padronizar_emissor_lf(emissor) -> str | None:
    if esta_vazio(emissor):
        return None

    emissor_original = str(emissor).strip()
    chave = normalizar_emissor_para_regra(emissor_original)

    return PADRONIZACAO_EMISSORES_LF.get(chave, emissor_original)


def abreviar_mes_pt_br(dt: datetime) -> str:
    meses = {
        1: "jan", 2: "fev", 3: "mar", 4: "abr",
        5: "mai", 6: "jun", 7: "jul", 8: "ago",
        9: "set", 10: "out", 11: "nov", 12: "dez",
    }
    return f"{meses[dt.month]}/{str(dt.year)[-2:]}"


def parse_data_segura(valor):
    if valor is None:
        return None

    if isinstance(valor, datetime):
        return valor

    if isinstance(valor, date):
        return datetime(valor.year, valor.month, valor.day)

    texto = str(valor).strip()
    if not texto or texto in {"-", "None", "nan", "NaN"}:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(texto, fmt)
        except Exception:
            continue

    try:
        return datetime.fromisoformat(texto)
    except Exception:
        return None


def gerar_codigo_let_financeira(emissor, vencimento) -> str | None:
    if esta_vazio(emissor) or esta_vazio(vencimento):
        return None

    dt = parse_data_segura(vencimento)
    if dt is None:
        return None

    emissor_padronizado = padronizar_emissor_lf(emissor)
    if not emissor_padronizado:
        return None

    return f"LF {emissor_padronizado} {abreviar_mes_pt_br(dt)}"

def gerar_codigo_padronizado_credito(prefixo: str, emissor, vencimento) -> str | None:
    if esta_vazio(emissor) or esta_vazio(vencimento):
        return None

    dt = parse_data_segura(vencimento)
    if dt is None:
        return None

    emissor_padronizado = padronizar_emissor_lf(emissor)
    if not emissor_padronizado:
        return None

    return f"{prefixo} {emissor_padronizado} {abreviar_mes_pt_br(dt)}"

def gerar_codigo_titulo_publico(sigla: str, vencimento) -> str | None:
    if esta_vazio(vencimento):
        return None

    dt = parse_data_segura(vencimento)
    if dt is None:
        return None

    return f"{sigla} {abreviar_mes_pt_br(dt)}"


# =========================================================
# LEITURA / PRESERVAÇÃO DA BASE MANUAL
# =========================================================
def carregar_base_manual_existente() -> Dict[str, dict]:
    if not ARQUIVO_BASE_MANUAL.exists():
        return {}

    wb = load_workbook(ARQUIVO_BASE_MANUAL)

    if ABA_BASE_MANUAL not in wb.sheetnames:
        return {}

    ws = wb[ABA_BASE_MANUAL]
    mapa = localizar_colunas(ws)
    if "Id_Ativo" not in mapa:
        return {}

    # lê aba oculta de controle, se existir
    controle_por_id: Dict[str, dict] = {}
    if ABA_CONTROLE_CHAVES in wb.sheetnames:
        ws_ctrl = wb[ABA_CONTROLE_CHAVES]
        mapa_ctrl = localizar_colunas(ws_ctrl)

        if all(col in mapa_ctrl for col in COLUNAS_CONTROLE_CHAVES):
            for row_idx in range(2, ws_ctrl.max_row + 1):
                id_ativo = normalizar_texto(ws_ctrl.cell(row=row_idx, column=mapa_ctrl["Id_Ativo"]).value)
                if id_ativo == "-":
                    continue

                controle_por_id[id_ativo] = {
                    "Chave_Original": normalizar_texto(
                        ws_ctrl.cell(row=row_idx, column=mapa_ctrl["Chave_Original"]).value
                    ),
                    "Chave_Final": normalizar_texto(
                        ws_ctrl.cell(row=row_idx, column=mapa_ctrl["Chave_Final"]).value
                    ),
                }

    base: Dict[str, dict] = {}

    for row_idx in range(2, ws.max_row + 1):
        id_ativo = normalizar_texto(ws.cell(row=row_idx, column=mapa["Id_Ativo"]).value)
        if id_ativo == "-":
            continue

        registro_esquerda = {}
        for col in COLUNAS_FIXAS:
            registro_esquerda[col] = formatar_valor_para_excel(
                col,
                ws.cell(row=row_idx, column=mapa[col]).value if col in mapa else "-",
            )

        registro_manual = {}
        for col_fixa, col_manual in zip(COLUNAS_FIXAS, COLUNAS_MANUAIS):
            if col_manual not in mapa:
                registro_manual[col_fixa] = "-"
                continue
            registro_manual[col_fixa] = formatar_valor_para_excel(
                col_fixa,
                ws.cell(row=row_idx, column=mapa[col_manual]).value,
            )

        controle = controle_por_id.get(id_ativo, {})
        chave_original = controle.get("Chave_Original")
        chave_final = controle.get("Chave_Final")

        if esta_vazio(chave_original):
            chave_original = gerar_chave_exata_ativo(registro_esquerda)

        if esta_vazio(chave_final):
            chave_final = gerar_chave_final_manual(registro_esquerda, registro_manual)

        base[id_ativo] = {
            "Id_Ativo": id_ativo,
            "Chave_Original": chave_original,
            "Chave_Final": chave_final,
            "Campos_Faltantes": normalizar_texto(
                ws.cell(row=row_idx, column=mapa["Campos_Faltantes"]).value
                if "Campos_Faltantes" in mapa
                else "-"
            ),
            "Registro": registro_esquerda,
            "Manual": registro_manual,
        }

    return base


def carregar_preenchimentos_manuais(base_existente: Dict[str, dict]) -> Dict[str, Dict[str, object]]:
    preenchimentos: Dict[str, Dict[str, object]] = {}

    for _, item in base_existente.items():
        dados: Dict[str, str] = {}
        for col in COLUNAS_FIXAS:
            valor_manual = item["Manual"].get(col)
            if not esta_vazio(valor_manual):
                dados[col] = formatar_valor_para_excel(col, valor_manual)

        chave_original = item.get("Chave_Original")
        chave_final = item.get("Chave_Final")

        if dados and chave_original and chave_original != "-":
            preenchimentos[chave_original] = {
                "dados": dados,
                "chave_final": chave_final,
            }

    return preenchimentos


# =========================================================
# TRATAMENTO DA CARTEIRA
# =========================================================
def aplicar_preenchimentos_na_carteira(wb_carteira, preenchimentos_manuais: Dict[str, Dict[str, object]]) -> int:
    total_alteracoes = 0

    for nome_aba in wb_carteira.sheetnames:
        ws = wb_carteira[nome_aba]
        mapa_colunas = localizar_colunas(ws)

        if not all(col in mapa_colunas for col in COLUNAS_FIXAS):
            print(f"[AVISO] A aba '{nome_aba}' não tem o cabeçalho esperado. Pulando.")
            continue

        for row_idx in iterar_linhas_dados(ws):
            valores_atuais = {
                col: ws.cell(row=row_idx, column=mapa_colunas[col]).value
                for col in COLUNAS_FIXAS
            }

            chave_original = gerar_chave_exata_ativo(valores_atuais)
            item_manual = preenchimentos_manuais.get(chave_original, {})
            dados_manuais = item_manual.get("dados", {})

            houve_alteracao = False
            if dados_manuais:
                for col in COLUNAS_FIXAS:
                    if col not in dados_manuais:
                        continue

                    col_idx = mapa_colunas[col]
                    valor_atual = ws.cell(row=row_idx, column=col_idx).value
                    novo_valor = dados_manuais[col]

                    if deve_aplicar_valor(valor_atual, novo_valor):
                        ws.cell(row=row_idx, column=col_idx, value=formatar_valor_para_excel(col, novo_valor))
                        if col == "Vencimento":
                            ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"
                        houve_alteracao = True

            if houve_alteracao:
                total_alteracoes += 1

    return total_alteracoes


def deduplicar_linhas_apos_preenchimento(wb_carteira) -> int:
    total_linhas_removidas = 0

    for nome_aba in wb_carteira.sheetnames:
        ws = wb_carteira[nome_aba]
        mapa = localizar_colunas(ws)

        if not all(col in mapa for col in COLUNAS_FIXAS):
            continue

        linhas = list(iterar_linhas_dados(ws))
        if not linhas:
            continue

        colunas_datas = []
        for col_idx in range(1, ws.max_column + 1):
            cab = ws.cell(row=1, column=col_idx).value
            if cab is None:
                continue
            cab = str(cab).strip()
            if cab in COLUNAS_FIXAS:
                continue
            if eh_coluna_data(cab):
                colunas_datas.append(col_idx)

        grupos: Dict[str, List[int]] = {}
        for row_idx in linhas:
            valores_fixos = {col: ws.cell(row=row_idx, column=mapa[col]).value for col in COLUNAS_FIXAS}
            chave_final = gerar_chave_exata_ativo(valores_fixos)
            grupos.setdefault(chave_final, []).append(row_idx)

        linhas_excluir: List[int] = []

        for _, grupo in grupos.items():
            if len(grupo) <= 1:
                continue

            linha_base = grupo[0]
            for linha_dup in grupo[1:]:
                for col in COLUNAS_FIXAS:
                    col_idx = mapa[col]
                    valor_base = ws.cell(row=linha_base, column=col_idx).value
                    valor_dup = ws.cell(row=linha_dup, column=col_idx).value

                    if esta_vazio(valor_base) and not esta_vazio(valor_dup):
                        ws.cell(row=linha_base, column=col_idx, value=formatar_valor_para_excel(col, valor_dup))
                        if col == "Vencimento":
                            ws.cell(row=linha_base, column=col_idx).number_format = "yyyy-mm-dd"

                for col_idx in colunas_datas:
                    valor_base = to_float_seguro(ws.cell(row=linha_base, column=col_idx).value)
                    valor_dup = to_float_seguro(ws.cell(row=linha_dup, column=col_idx).value)
                    soma = valor_base + valor_dup
                    ws.cell(row=linha_base, column=col_idx, value=soma)
                    ws.cell(row=linha_base, column=col_idx).number_format = "#,##0.00"

                linhas_excluir.append(linha_dup)

        for row_idx in sorted(set(linhas_excluir), reverse=True):
            ws.delete_rows(row_idx, 1)
            total_linhas_removidas += 1

    return total_linhas_removidas


def coletar_ativos_unicos_da_carteira(wb_carteira) -> Dict[str, dict]:
    ativos_unicos: Dict[str, dict] = {}

    for nome_aba in wb_carteira.sheetnames:
        ws = wb_carteira[nome_aba]
        mapa = localizar_colunas(ws)
        if not all(col in mapa for col in COLUNAS_FIXAS):
            continue

        for row_idx in iterar_linhas_dados(ws):
            valores = {col: ws.cell(row=row_idx, column=mapa[col]).value for col in COLUNAS_FIXAS}
            registro = normalizar_registro_fixos(valores)

            chave_original = gerar_chave_exata_ativo(registro)
            id_ativo = gerar_id_por_chave(chave_original)
            campos_faltantes = contar_faltantes(registro)

            if id_ativo not in ativos_unicos:
                ativos_unicos[id_ativo] = {
                    "Id_Ativo": id_ativo,
                    "Chave_Original": chave_original,
                    "Chave_Final": chave_original,
                    "Campos_Faltantes": ", ".join(campos_faltantes) if campos_faltantes else "-",
                    "Registro": registro,
                    "Manual": registro.copy(),
                }
            else:
                existente = ativos_unicos[id_ativo]["Registro"]
                houve_melhoria = False
                for col in COLUNAS_FIXAS:
                    if esta_vazio(existente.get(col)) and not esta_vazio(registro.get(col)):
                        existente[col] = registro[col]
                        houve_melhoria = True

                if houve_melhoria:
                    campos = contar_faltantes(existente)
                    ativos_unicos[id_ativo]["Campos_Faltantes"] = ", ".join(campos) if campos else "-"
                    ativos_unicos[id_ativo]["Manual"] = existente.copy()
                    ativos_unicos[id_ativo]["Chave_Final"] = gerar_chave_exata_ativo(existente)

    return ativos_unicos


# =========================================================
# GERAÇÃO DA BASE MANUAL AUXILIAR
# =========================================================
def ajustar_larguras_base_manual(ws) -> None:
    larguras = {
        "A": 24,  # Id_Ativo
        "B": 22,  # Campos_Faltantes
        "C": 16,  # Codigo
        "D": 20,  # ISIN
        "E": 28,  # Descrição
        "F": 24,  # Emissor
        "G": 22,  # Classe
        "H": 26,  # Tipo de Investimento
        "I": 16,  # Vencimento
        "J": 16,  # Codigo - Preencher
        "K": 20,  # ISIN - Preencher
        "L": 28,  # Descrição - Preencher
        "M": 24,  # Emissor - Preencher
        "N": 22,  # Classe - Preencher
        "O": 26,  # Tipo de Investimento - Preencher
        "P": 16,  # Vencimento - Preencher
    }

    for letra, largura in larguras.items():
        ws.column_dimensions[letra].width = largura

    for col_idx in range(1, ws.max_column + 1):
        letra = get_column_letter(col_idx)
        if letra not in larguras:
            ws.column_dimensions[letra].width = 20


def montar_base_manual_final(base_existente: Dict[str, dict], ativos_atuais: Dict[str, dict]) -> Dict[str, dict]:
    base_final: Dict[str, dict] = {}

    antigos_por_chave_final: Dict[str, dict] = {}

    for _, item in base_existente.items():
        registro_antigo = item.get("Registro", {}).copy()
        manual_antigo = item.get("Manual", {}).copy()

        chave_final_antiga = gerar_chave_final_manual(registro_antigo, manual_antigo)

        item_existente = antigos_por_chave_final.get(chave_final_antiga)
        if item_existente is None:
            antigos_por_chave_final[chave_final_antiga] = item
        else:
            score_atual = contar_preenchidos(item_existente.get("Manual", {}))
            score_novo = contar_preenchidos(manual_antigo)
            if score_novo > score_atual:
                antigos_por_chave_final[chave_final_antiga] = item

    for id_ativo_atual, item_atual in ativos_atuais.items():
        registro_atual = item_atual["Registro"].copy()
        chave_atual = gerar_chave_exata_ativo(registro_atual)

        item_antigo = antigos_por_chave_final.get(chave_atual)

        if item_antigo:
            manual_final = item_antigo.get("Manual", {}).copy()

            for col in COLUNAS_FIXAS:
                if esta_vazio(manual_final.get(col)):
                    manual_final[col] = registro_atual.get(col, "-")
        else:
            manual_final = registro_atual.copy()

        classe = normalizar_texto(registro_atual.get("Classe")).strip().lower()
        codigo_original = normalizar_texto(registro_atual.get("Codigo")).strip().lower()

        codigo_automatico = None

        if classe == "letra financeira":
            codigo_automatico = gerar_codigo_padronizado_credito(
                "LF",
                registro_atual.get("Emissor"),
                registro_atual.get("Vencimento"),
            )

        elif classe == "cdb/ rdb":
            codigo_automatico = gerar_codigo_padronizado_credito(
                "CDB/RDB",
                registro_atual.get("Emissor"),
                registro_atual.get("Vencimento"),
            )

        elif classe == "dpge":
            codigo_automatico = gerar_codigo_padronizado_credito(
                "DPGE",
                registro_atual.get("Emissor"),
                registro_atual.get("Vencimento"),
            )

        elif codigo_original == "letras do tesouro nacional":
            codigo_automatico = gerar_codigo_titulo_publico(
                "LTN",
                registro_atual.get("Vencimento"),
            )

        elif codigo_original == "letras financeiras do tesouro":
            codigo_automatico = gerar_codigo_titulo_publico(
                "LFT",
                registro_atual.get("Vencimento"),
            )

        elif codigo_original == "notas do tesouro nacional serie b":
            codigo_automatico = gerar_codigo_titulo_publico(
                "NTN-B",
                registro_atual.get("Vencimento"),
            )

        elif codigo_original == "notas do tesouro nacional - serie f":
            codigo_automatico = gerar_codigo_titulo_publico(
                "NTN-F",
                registro_atual.get("Vencimento"),
            )

        if codigo_automatico:
            manual_final["Codigo"] = codigo_automatico

        base_final[id_ativo_atual] = {
            "Id_Ativo": id_ativo_atual,
            "Chave_Original": item_antigo.get("Chave_Original", chave_atual) if item_antigo else chave_atual,
            "Chave_Final": gerar_chave_final_manual(registro_atual, manual_final),
            "Campos_Faltantes": item_atual["Campos_Faltantes"],
            "Registro": registro_atual,
            "Manual": manual_final,
        }

    return base_final


def gerar_base_manual(base_final: Dict[str, dict]) -> None:
    wb = Workbook()

    # =========================
    # ABA VISÍVEL
    # =========================
    ws = wb.active
    ws.title = ABA_BASE_MANUAL
    ws.append(COLUNAS_BASE_MANUAL)

    fill_header_esquerda = PatternFill(fill_type="solid", fgColor="1F4E78")
    fill_header_direita = PatternFill(fill_type="solid", fgColor="548235")
    fill_highlight = PatternFill(fill_type="solid", fgColor="FFF2CC")
    fill_divisoria = PatternFill(fill_type="solid", fgColor="D9E2F3")

    for col_idx, titulo in enumerate(COLUNAS_BASE_MANUAL, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if titulo in COLUNAS_MANUAIS:
            cell.fill = fill_header_direita
        else:
            cell.fill = fill_header_esquerda

    linhas_ordenadas = sorted(
        base_final.values(),
        key=lambda item: (
            normalizar_texto(item["Registro"].get("Descrição")),
            normalizar_texto(item["Registro"].get("Codigo")),
            item["Id_Ativo"],
        ),
    )

    for item in linhas_ordenadas:
        registro = item["Registro"]
        manual = item["Manual"]

        linha = [
            item["Id_Ativo"],
            item["Campos_Faltantes"],
        ]
        linha.extend(registro.get(col, "-") for col in COLUNAS_FIXAS)
        linha.extend(manual.get(col, "-") for col in COLUNAS_FIXAS)
        ws.append(linha)

    mapa = localizar_colunas(ws)
    idx_inicio_manual = mapa[COLUNAS_MANUAIS[0]]

    for row_idx in range(2, ws.max_row + 1):
        campos_faltantes = normalizar_texto(ws.cell(row=row_idx, column=mapa["Campos_Faltantes"]).value)
        faltantes_set = {campo.strip() for campo in campos_faltantes.split(",") if campo.strip() and campo.strip() != "-"}

        for col in COLUNAS_FIXAS:
            col_esquerda = mapa[col]
            col_direita = mapa[f"{col} - Preencher"]

            ws.cell(row=row_idx, column=col_esquerda).alignment = Alignment(vertical="center")
            ws.cell(row=row_idx, column=col_direita).alignment = Alignment(vertical="center")

            if col == "Vencimento":
                ws.cell(row=row_idx, column=col_esquerda).number_format = "yyyy-mm-dd"
                ws.cell(row=row_idx, column=col_direita).number_format = "yyyy-mm-dd"

            if col in faltantes_set:
                ws.cell(row=row_idx, column=col_esquerda).fill = fill_highlight
                ws.cell(row=row_idx, column=col_direita).fill = fill_highlight

        ws.cell(row=row_idx, column=idx_inicio_manual - 1).fill = fill_divisoria

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ajustar_larguras_base_manual(ws)

    # =========================
    # ABA OCULTA DE CONTROLE
    # =========================
    ws_ctrl = wb.create_sheet(ABA_CONTROLE_CHAVES)
    ws_ctrl.append(COLUNAS_CONTROLE_CHAVES)

    for item in linhas_ordenadas:
        ws_ctrl.append([
            item["Id_Ativo"],
            item.get("Chave_Original", "-"),
            item.get("Chave_Final", "-"),
        ])

    ws_ctrl.sheet_state = "hidden"

    wb.save(ARQUIVO_BASE_MANUAL)


# =========================================================
# PÓS-TRATAMENTO
# =========================================================
def formatar_datas_em_todas_as_abas(wb) -> None:
    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        mapa = localizar_colunas(ws)
        if "Vencimento" not in mapa:
            continue

        col_venc = mapa["Vencimento"]

        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_venc)
            if esta_vazio(cell.value):
                continue

            cell.value = formatar_valor_para_excel("Vencimento", cell.value)
            cell.number_format = "yyyy-mm-dd"


def executar_tratamento() -> None:
    if not ARQUIVO_CARTEIRA.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {ARQUIVO_CARTEIRA}")

    base_existente = carregar_base_manual_existente()
    preenchimentos_manuais = carregar_preenchimentos_manuais(base_existente)
    wb_carteira = load_workbook(ARQUIVO_CARTEIRA)

    total_alteracoes = aplicar_preenchimentos_na_carteira(
        wb_carteira,
        preenchimentos_manuais,
    )

    total_linhas_removidas = deduplicar_linhas_apos_preenchimento(wb_carteira)
    formatar_datas_em_todas_as_abas(wb_carteira)

    ativos_atuais = coletar_ativos_unicos_da_carteira(wb_carteira)
    base_final = montar_base_manual_final(base_existente, ativos_atuais)

    wb_carteira.save(ARQUIVO_SAIDA)
    gerar_base_manual(base_final)

    print(f"Arquivo tratado com sucesso: {ARQUIVO_SAIDA}")
    print(f"Base auxiliar gerada/atualizada: {ARQUIVO_BASE_MANUAL}")
    print(f"Total de linhas alteradas na carteira: {total_alteracoes}")
    print(f"Total de linhas duplicadas removidas: {total_linhas_removidas}")
    print(f"Total de chaves na base auxiliar: {len(base_final)}")
    print(f"Total de ativos atuais únicos na carteira: {len(ativos_atuais)}")


if __name__ == "__main__":
    executar_tratamento()
