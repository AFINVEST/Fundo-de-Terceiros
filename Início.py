from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Dict, Tuple
from textwrap import dedent
from html import escape

import openpyxl
import streamlit as st
import streamlit.components.v1 as components
from openpyxl.utils import get_column_letter, range_boundaries
from ui_padrao import aplicar_ui_padrao

# Para atualizar o site, deve rodar no vs code:

# git add .gitignore
# git add cadastro_manual_ativos.xlsx
# git add carteira_fundos_consolidada.xlsx
# git add consolidar_planilha.py
# git add tratar_planilha_consolidada.py
# git add baixar_carteiras_cvm.py

# Depois:
# git status

# Depois: 
# git commit -m "Atualiza consolidação das carteiras dos fundos"

# Depois:
# git push

st.set_page_config(page_title="Comparador de Fundos", layout="wide")
aplicar_ui_padrao()
TITULO_FIXO = "Comparador de Fundos"

# BASE_DIR = Path(r"Z:\Asset Management\Equipe\Lívia\Rascunhos\Site_fundos")
BASE_DIR = Path(__file__).resolve().parent
ARQUIVO_EXCEL = BASE_DIR / "Capa e correlação - site.xlsx"
ABA_CAPA = "CAPA"

COL_INICIAL = 2   # B
COL_FINAL = 12    # L

MESES_PT = {
    1: "jan",
    2: "fev",
    3: "mar",
    4: "abr",
    5: "mai",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "set",
    10: "out",
    11: "nov",
    12: "dez",
}


def _carregar_planilha(caminho: Path):
    wb_estilo = openpyxl.load_workbook(caminho, data_only=False)
    wb_valor = openpyxl.load_workbook(caminho, data_only=True)
    return wb_estilo[ABA_CAPA], wb_valor[ABA_CAPA]


def _montar_intervalo_dinamico(ws) -> str:
    return f"{get_column_letter(COL_INICIAL)}2:{get_column_letter(COL_FINAL)}{ws.max_row}"


def _mapa_merged_cells(
    ws, intervalo: str
) -> Tuple[Dict[Tuple[int, int], Tuple[int, int]], set[Tuple[int, int]]]:
    min_col, min_row, max_col, max_row = range_boundaries(intervalo)
    merges: Dict[Tuple[int, int], Tuple[int, int]] = {}
    ignorar: set[Tuple[int, int]] = set()

    for merged in ws.merged_cells.ranges:
        c1, r1, c2, r2 = merged.bounds
        if r2 < min_row or r1 > max_row or c2 < min_col or c1 > max_col:
            continue

        rowspan = r2 - r1 + 1
        colspan = c2 - c1 + 1
        merges[(r1, c1)] = (rowspan, colspan)

        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    ignorar.add((r, c))

    return merges, ignorar


def _formatar_numero_br(valor: float, casas: int = 2) -> str:
    texto = f"{valor:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def _formatar_valor(valor, numero_format: str) -> str:
    if valor is None:
        return ""

    numero_format = (numero_format or "General").lower()

    if isinstance(valor, datetime):
        if "mmm" in numero_format:
            return f"{MESES_PT[valor.month]}/{str(valor.year)[-2:]}"
        return valor.strftime("%d/%m/%Y")

    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")

    if isinstance(valor, (int, float)):
        if "%" in numero_format:
            casas = 2 if ".00" in numero_format or "0,00" in numero_format else 1
            return f"{_formatar_numero_br(valor * 100, casas)}%"

        if "#,##0.00" in numero_format or "0.00" in numero_format:
            return _formatar_numero_br(valor, 2)

        if "0.0" in numero_format:
            return _formatar_numero_br(valor, 1)

        if isinstance(valor, float) and not valor.is_integer():
            return _formatar_numero_br(valor, 2)

        return _formatar_numero_br(valor, 0)

    return str(valor)


def _alinhamento_css(celula, valor) -> str:
    if celula.alignment and celula.alignment.horizontal:
        return celula.alignment.horizontal

    if isinstance(valor, (int, float, datetime, date)):
        return "right"

    return "left"


def _linha_vazia(ws_valor, linha: int, min_col: int, max_col: int) -> bool:
    for coluna in range(min_col, max_col + 1):
        if ws_valor.cell(linha, coluna).value not in (None, ""):
            return False
    return True


def _classe_linha(linha_excel: int, valor_b: str | None, linha_vazia: bool) -> str:
    if linha_vazia:
        return "espaco"
    if linha_excel == 2:
        return "titulo"
    if valor_b == "Dados":
        return "bloco"
    if valor_b == "Fundos D+1":
        return "subtitulo_d1"
    if valor_b == "Fundos D+30":
        return "subtitulo_d30"
    if valor_b == "Fundos D+60":
        return "subtitulo_d60"
    if valor_b == "Nome":
        return "cabecalho"
    return "dados"


def _estilo_inline(celula, valor, classe_linha: str, coluna: int, max_col: int) -> str:
    estilos = []

    alinhamento = _alinhamento_css(celula, valor)
    estilos.append(f"text-align:{alinhamento};")

    fonte = celula.font
    if fonte:
        if fonte.name:
            estilos.append(f"font-family:'{fonte.name}', 'Segoe UI', Arial, sans-serif;")
        if fonte.sz:
            estilos.append(f"font-size:{fonte.sz}pt;")
        estilos.append(f"font-weight:{700 if fonte.bold else 400};")
        if fonte.italic:
            estilos.append("font-style:italic;")
        if fonte.underline and fonte.underline != "none":
            estilos.append("text-decoration:underline;")

    # FORÇAR O TÍTULO
    if classe_linha == "titulo":
        estilos.append("font-size:20pt !important;")
        estilos.append("font-weight:700 !important;")
        estilos.append("line-height:1.2;")

    if alinhamento == "center":
        estilos.append("padding-left:4px; padding-right:4px;")

    if coluna == max_col and classe_linha == "dados":
        estilos.append("padding-right:6px;")

    if classe_linha == "espaco":
        estilos.append("border:none; padding:0; background:white;")

    return " ".join(estilos)


def renderizar_capa_excel(caminho_arquivo: Path) -> tuple[str, int]:
    ws_estilo, ws_valor = _carregar_planilha(caminho_arquivo)

    intervalo = _montar_intervalo_dinamico(ws_estilo)
    min_col, min_row, max_col, max_row = range_boundaries(intervalo)
    merges, ignorar = _mapa_merged_cells(ws_estilo, intervalo)

    larguras = []
    for coluna in range(min_col, max_col + 1):
        letra = get_column_letter(coluna)
        largura_excel = ws_estilo.column_dimensions[letra].width or 12
        largura_px = max(int(largura_excel * 7), 90)
        larguras.append(f"<col style='width:{largura_px}px'>")

    html = [dedent("""
        <style>
            body {
                margin: 0;
                padding: 0;
                background: white;
            }
            .excel-wrap {
                overflow-x: auto;
                padding: 8px 12px;
                background: white;
            }
            table.excel-like {
                border-collapse: collapse;
                table-layout: fixed;
                width: max-content;
                min-width: 1450px;
                background: white;
                color: #000;
                font-family: Calibri, 'Segoe UI', Arial, sans-serif;
                font-size: 11pt;
                line-height: 1.15;
            }
            .excel-like td {
                border: 1px solid #1f1f1f;
                padding: 2px 4px;
                vertical-align: middle;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }
            .excel-like td.titulo {
                border: none !important;
                padding: 6px 0 18px 0;
            }
            .excel-like td.instrucao,
            .excel-like td.bloco {
                border: none !important;
                padding: 3px 0;
            }
            .excel-like td.espaco {
                border: none !important;
                padding: 0 !important;
                height: 10px !important;
                background: white !important;
            }
            .excel-like td.subtitulo_d1 {
                background: #D9E1F2;
                font-weight: 700 !important;
            }

            .excel-like td.subtitulo_d30 {
                background: #9AB2DE;
                font-weight: 700 !important;
            }

            .excel-like td.subtitulo_d60 {
                background: #678CCF;
                font-weight: 700 !important;
            }
                   
            .excel-like td.cabecalho {
                font-weight: 700 !important;
            }    
        </style>
        <div class="excel-wrap">
        <table class="excel-like">
        <colgroup>
    """)]

    html.extend(larguras)
    html.append("</colgroup>")

    for linha in range(min_row, max_row + 1):
        if linha == 2:
            continue
        linha_vazia = _linha_vazia(ws_valor, linha, min_col, max_col)
        valor_b = ws_valor.cell(linha, min_col).value
        classe_linha = _classe_linha(linha, valor_b, linha_vazia)

        altura = ws_estilo.row_dimensions[linha].height
        if classe_linha == "espaco":
            altura_css = "height:10px;"
        elif altura:
            altura_css = f"height:{int(altura * 1.33)}px;"
        else:
            altura_css = ""

        html.append(f"<tr style='{altura_css}'>")

        for coluna in range(min_col, max_col + 1):
            if (linha, coluna) in ignorar:
                continue

            celula_estilo = ws_estilo.cell(linha, coluna)
            celula_valor = ws_valor.cell(linha, coluna)
            valor = celula_valor.value

            rowspan, colspan = merges.get((linha, coluna), (1, 1))
            attrs = []

            if rowspan > 1:
                attrs.append(f"rowspan='{rowspan}'")
            if colspan > 1:
                attrs.append(f"colspan='{colspan}'")

            conteudo = escape(_formatar_valor(valor, celula_estilo.number_format))
            estilo_inline = _estilo_inline(celula_estilo, valor, classe_linha, coluna, max_col)

            attrs.append(f"class='{classe_linha}'")
            attrs.append(f"style='{estilo_inline}'")

            html.append(f"<td {' '.join(attrs)}>{conteudo}</td>")

        html.append("</tr>")

    html.append("</table></div>")
    return "\n".join(html), max_row


if not ARQUIVO_EXCEL.exists():
    st.error(f"Não encontrei o arquivo: {ARQUIVO_EXCEL}")
    st.stop()

st.markdown(
    f"""
    <div style="
        font-family: Calibri, 'Segoe UI', Arial, sans-serif;
        font-size: 40px;
        font-weight: 700;
        color: black;
        margin-bottom: 12px;
    ">
        {TITULO_FIXO}
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown("""
<div style="font-family: Calibri, 'Segoe UI', Arial, sans-serif; line-height: 1.2; margin-top: 0;">
    <p style="margin: 0 0 6px 0;">Use o menu lateral para abrir:</p>
    <p style="margin: 0 0 2px 0;">1- Fundos D+1</p>
    <p style="margin: 0 0 2px 0;">2- Fundos D+30</p>
    <p style="margin: 0 0 2px 0;">3- Fundos D+60</p>
    <p style="margin: 0;">4- Análise Fundo</p>
</div>
""", unsafe_allow_html=True)

html, max_row = renderizar_capa_excel(ARQUIVO_EXCEL)
altura = max(1100, int(max_row * 32))

components.html(
    html,
    height=altura,
    scrolling=False,
)
